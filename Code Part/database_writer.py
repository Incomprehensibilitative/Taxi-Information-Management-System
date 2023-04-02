import openpyxl

data = openpyxl.load_workbook("Taxi-information.xlsx", data_only=True)

def write_data(system):
    # rewrite data for "Customer" sheet:
    sheet1 = data['Customer']
    sheet1.delete_rows(2, sheet1.max_row-1)
    for cus in system.get_list("customer"):
        sheet1.append((cus.get_id(), cus.get_name(), cus.get_phone_num(), cus.get_chosen_vehicle()))
    
    # rewrite data for "Driver" sheet:
    sheet2 = data['Driver']
    sheet2.delete_rows(2, sheet2.max_row-1)
    for drv in system.get_list("driver"):
        sheet2.append((drv.get_id(), drv.get_name(), drv.get_phone_num(), drv.get_vehicle_id(), drv.get_salary(), drv.get_gender(), drv.get_age()))

    # rewrite data for "Vehicle" sheet:
    sheet3 = data['Vehicle']
    sheet3.delete_rows(2, sheet3.max_row-1)
    for vhc in system.get_list("vehicle"):
        sheet3.append((vhc.get_id(), vhc.get_type(), vhc.get_regis_num(), vhc.get_price()))

    sheet4 = data["Invoice"]
    sheet4.delete_rows(2, sheet4.max_row-1)
    for inv in system.get_list("invoice"):
        sheet4.append((inv.get_id(), inv.get_customer_id(), inv.get_driver_id(), inv.get_date(), inv.get_payment_mode(), inv.get_distance(), inv.get_price_per_km(), inv.get_total()))

    data.save("Taxi-information.xlsx")